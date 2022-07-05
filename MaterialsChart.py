# # reading PDF
import os
import sys
from tkinter import filedialog

# may want to use pdfplumber and my own parser, instead of convertin to xlsx
# because sometimes line have two lines worth of data in a single cell
# Also, should try out different pdf to xlsx converters bluebeam might not be consistent


# import pdfplumber 
    

# acronyms = ["CHWR", "CHWS", "COND DRAIN", "DCW", "DHW", Gas, HHWR, HHWS, 
#     SW UG, SW AG, STORM AG, STORM UG, STORM, ]    
# meanings = [Chill Water, Chill Water, Condensate Drain, Domestic Cold Water,
#     Domestic Hot Water, Natural Gas, Heating Hot Water, Heating Hot Water,
#     Waste Underground, Waste Above Ground, Storm Above Ground, Storm Under Ground,
#     Storm ANSI, 
# with pdfplumber.open(r'ChilledWater.pdf') as pdf:
#      first_page = pdf.pages[0]
#      print(first_page.extract_text())


# from asyncio import subprocess
# from itertools import count
# import sys
# from tracemalloc import stop
# from typing import Iterator
# from unicodedata import name
import openpyxl
import re
import pathos as p
import subprocess
from fractions import Fraction

systems = {
    "CHWR": "Chill Water Return",
    "CHWS": "Chill Water Supply",
    "COND DRAIN": "Condensate Drain",
    "DCW": "Domestic Cold Water",
    "DHW": "Domestic Hot Water",
    # "Gas": "Natural Gas", #Make sure on this one, and no false positives
    "HHWR": "Heating Hot Water Return",
    "HHWS": "Heating Hot Water Supply",
    "SW UG": "Waste Underground",
    "SW AG": "Waste Above Ground",
    "STORM AG": "Storm Above Ground",
    "STORM UG": "Storm Under Ground",
    "STORM": "Storm ANSI" #Not sure about this one
    # "HUSKY", might want to check for this
}

#make function for checking copper type, if need be
pipeMaterials = {
    "Carbon Steel": "SCH.40 Steel",
    "Copper": "Copper",
    "Cast Iron": "Cast Iron",
    # Use regex to allow anything in ****
    # Not sure if there is ever a situation where the PVC is not solid core.
    "PVC - **** Sch 40": "SOLID CORE SCH.40 PVC"
}

copperTypes = [ 
    "TYPE L", "TYPE K", "PRESSFIT"
   ]

connectionTypes = {
    "ButtWld": "BUTT-WELD",
    "Pressure": "PRESS JOINT",
    "PressFit": "PRESS JOINT",
    "ProPress": "PRESS JOINT",
    "NoHub": "HUSKY 2000",
    "SolvWld": "Solvent",
    "Thread": "Threaded"
}

locations = {
    "AG": "ABOVE GROUND",
    "UG": "UNDERGROUND"
}



class Section:
    # might be better to use a dictionary instead with
    sectionNames = ["Recap", "Pipe", "Nipples", "Fittings"]
    # may or may not want this default value
    def __init__(self, name = "", start = 0, end = 0):
        self.name = name
        self.start = start
        self.end = end
    def __str__(self):
        return (self.name + " " + str(self.start) + "," + str(self.end))
    def __repr__(self):
        return self.__str__()

class PipeSize:
    def __init__(self, i = 0, f = Fraction(0)):
        self.i = i
        self.f = f
    def __str__(self):
        return str(self.i) + str(self.f)
    def __repr__(self):
        return self.__str__()

class Fitting:
    fittingMaterials = {
        #unsure if it's ever just Copper or just Bronze
        "Copper": "Copper/Bronze",
        "Bronze": "Copper/Bronze",
        "NoHub": "No-Hub",
        # not sure if their are other Malleable Iron numbers
        # use regex
        "Malleable Iron **** Class 150": "150# Malleable Iron",
        # is 3000 vs 2000 important? Specs seem to say 3000
        # "Carbon Steel - Branch Outlets 3000# - ButtWld"
        # Might want to use regex
        "Carbon Steel - Branch Outlets ****#": "2000# Forged Steel"
    }
    def __init__(self, material = "", connectionType = "", manufacturer = ""):
        self.material = material
        self.connectionType = connectionType
        self.manufacturer = manufacturer
        # Force material, connectionType and  manufacterer to be in the lists above
    def __str__(self):
        return("{} {} {}".format(self.material, self.connectionType, self.manufacturer))# use string formatting
    def __repr__(self):
        return(self.__str__())

activePipeMaterials = []    
sections: list[Section] = []

sourcePath = "ChilledWater.xlsx"
finalPath = "final.xlsx"



def findPipeMaterials(iterator):
    counter = 1
    # Go through row one until material is found (Copper, Steel, Etc)
    for row in iterator:
        # may need to use regex
        for m in pipeMaterials:
            if(row[0].value is not None and m in str(row[0].value)):
                #should figure out how to remove this loop
                for s in sections:
                    if(s.name == "Pipe"):
                        if(row[0].row > s.end):
                            return
                finalSheet.cell(column=3+currentColumn, row=counter).value = row[0].value
                #rowIter = sourceSheet.iter_rows(min_row=counter, max_col=sourceSheet.max_column)
                print(row[0].value)
                #findPipeQuantity(rowIter)
                #finalSheet.cell(column=7+currentColumn, row=counter).value = m
                activePipeMaterials.append(m)
                findPipeSizes(iterator, counter)
                counter += 1

def findPipeQuantity(iterator):
    pipeQuantity = []
    readingQuantity = False
    totalInches = 0
    quantityColumn = 0
    for row in iterator:
        #print(row[6].value)
        if(readingQuantity):
            #this isn't perfect because sometimes their is a blank line below quantity
            if(isinstance(row[quantityColumn].value, int)):
                totalInches += row[quantityColumn].value
                print("Total: {}".format(totalInches))
            else:
                #print(totalInches)
                pipeQuantity.append((totalInches))
                readingQuantity = False
                totalInches = 0
                print("PipeQuantity: " + str(pipeQuantity))
                return
        for c in row:
            if(c.value is not None and isinstance(c.value, str) and "quantity" in str(c.value).lower()):
                # print(c.column)
                quantityColumn = c.column - 1
                readingQuantity = True
                break
    print(pipeQuantity)    
            
    print("{} total inches".format(totalInches))    

def findPipeSizes(iterator, counter):
    # we find item and size, in the first row,
    # Scan from first line without letters to last line without letters
    started = False
    sizes = []
    for row in iterator:
        r0Value = row[0].value
        #print(str(row[6].value))
        if(r0Value is not None):
            # this doesn't always work, sometimes a cell has text then a number on the line below
            # and thus isn't an int type
            if(isinstance(r0Value, int)):
                sizes.append(str(r0Value))
                started = True
            elif(isinstance(r0Value, str) and '\n' in r0Value):
                #parse out the real value, probably need a try catch, or with
                pass
            elif(isinstance(r0Value, str) and not re.search('[a-zA-Z]', row[0].value)):
                #might want to add to the regex above instead of this if below
                if(not " " in r0Value):
                    sizes.append((r0Value))
                started = True
            elif(started):
                break
    print(sizes)
    if(len(sizes) > 1):
        finalSheet.cell(column=4+currentColumn, row=counter).value = sizes[0] + '" TO ' + sizes[-1] + '"'
    elif(len(sizes) > 0):
        finalSheet.cell(column=4+currentColumn, row=counter).value = sizes[0]

def findFittingMaterials(iterator, end):
    counter = 1
    # Go through row one until material is found (Copper, Steel, Etc)
    # Finds the end of the fittings section
    for s in sections:
        if(s.name == "Fittings"):
            finalRow = s.end
    #Why is pm not being used?
    for pm in activePipeMaterials:
        iterator = sourceSheet.iter_rows(max_row=end, max_col=sourceSheet.max_column)
        for row in iterator:
            for m in Fitting.fittingMaterials:
                if(row[0].value is not None and m in str(row[0].value)):
                    if(row[0].row > finalRow):
                        return
                    #finalSheet.cell(column=6+currentColumn, row=counter).value = row[0].value
                    #finalSheet.cell(column=8+currentColumn, row=counter).value = m
                    counter += 1

def recap(counter):
    iterator = sourceSheet.iter_rows(max_row=sourceSheet.max_row, max_col=sourceSheet.max_column)  
    counter = 1
    #Maybe search for keyword "System:" to make this safer and more efficient
    for row in iterator:
        for sys in systems:
            if(row[0].value is not None and sys in str(row[0].value)):
                # print(str(row[0].row) + systems[s])
                # might have both above and under ground in same PDF
                finalSheet.cell(column=5+currentColumn, row=counter).value = "UNDERGROUND" if sys + " UG" in str(row[0].value) else "ABOVE GROUND"
                # if(sys + " UG" in str(row[0].value)):
                #     finalSheet['F' + str(counter)].value = "UNDERGROUND"
                # else: "ABOVE GROUND"
                finalSheet.cell(column=1+currentColumn, row=counter).value = systems[sys]
                counter += 1  
    return counter                


filetypes = (
        ('Excel Workbook', '*.xlsx *.xlsm'),
        ('All files', '*.*')
    )


def readFile():
    # Iterates, finding sections
    counter = 1
    rowIter = sourceSheet.iter_rows(max_row=sourceSheet.max_row, max_col=sourceSheet.max_column)
    currentSection = Section()
    for row in rowIter:
        for sect in Section.sectionNames: 
            if(row[0].value is not None and sect in str(row[0].value)):
                if(not sect == currentSection.name):
                    if(currentSection.name == ""):
                        currentSection.start = row[0].row
                    else:
                        currentSection.end = row[0].row - 1
                        sections.append(currentSection)
                        finalSheet.cell(column=2+currentColumn, row=counter).value = str(currentSection)
                        counter += 1
                        currentSection = Section(start=row[0].row)
                currentSection.name = sect
    currentSection.end = row[0].row
    #finalSheet['B' + str(counter)].value = str(currentSection)
    sections.append(currentSection)            
    
    
     
    for sect in sections:
        # finds systems
        if(sect.name == "Recap"):
            recap(counter)
        #finds pipe materials
        if(sect.name == "Pipe"):
            rowIter = sourceSheet.iter_rows(max_row=sect.end, max_col=sourceSheet.max_column)
            findPipeMaterials(rowIter)   
        if(sect.name == "Fittings"):
            rowIter = sourceSheet.iter_rows(max_row=sect.end, max_col=sourceSheet.max_column)
            findFittingMaterials(rowIter, sect.end)

            
         
    
                
    # check that the last section has an endpoint
    #rowIter = sourceSheet.iter_rows(max_row=sect.end, max_col=sourceSheet.max_column)
    #findPipeQuantity(rowIter)
    print(sections)    
    

currentColumn = 0
if __name__ == '__main__':
    #main lives in the same namespace as top level code interesting, so no need for global calls
    final = openpyxl.Workbook()
    folderPath = filedialog.askdirectory(title="Select Folder With Material Charts")
    for file in os.listdir(folderPath):
        sourcePath  = os.path.join(folderPath, file)
        print(sourcePath)
        source = openpyxl.load_workbook(sourcePath)
        sourceSheet = source.active
        finalSheet = final.active
        readFile()
        currentColumn += 11
    final.save(filename = "final.xlsx")
    # Give the location of the file
    subprocess.run("explorer final.xlsx")
    