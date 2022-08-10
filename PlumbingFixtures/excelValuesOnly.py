from os import getcwd
import os.path as osp
from tkinter import filedialog
import openpyxl
import xlwings as xw
import sys
import time


def convert(filePath, sheetName):
    print("Converting page {} of {} to Values Only".format(sheetName, filePath))
    owb = openpyxl.load_workbook(filePath)
    ows = owb.active
    maxRow = ows.max_row
    maxCol = ows.max_column
    owb.close()
    exwb = xw.Book(filePath)
    exws = exwb.sheets[sheetName]
    exwb.app.calculate()
    #exws.range('G42').value = exws.range('G42').value
    #Row, Column
    exws.range((1,1),(10000, 100)).value = exws.range((1,1),(10000, 100)).value
    
    exwb.save()
    exwb.app.quit()
    print("Converted Successfully!")
    time.sleep(1)

def main():
    filePath = 'ToEvaluate/csm.xlsx'
    sheetName = 'Plbg'
    if(len(sys.argv) > 2):
        sheetName = sys.argv[2]
        filePath = sys.argv[1]
    else:
        filePath = filedialog.askopenfilename(title = "Select the file to convert")
        sheetName = input("Enter the sheet name: ")
    convert(filePath, sheetName)
    

if __name__ == '__main__':
    main()


