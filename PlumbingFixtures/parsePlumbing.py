from __future__ import annotations
import os
import subprocess
import sys
import tkinter
from typing import List
import openpyxl
from tkinter import filedialog
import excelValuesOnly as xl

#Figure out how to set the current working directory to the shortcut location
       
class plumbingJob:
    class partRow:
        def __init__(self, id, description: str, quantity, size_strength = None
          , manufacturer = "unknown", row = None, bidDayUnitPrice = None):
            self.row = row
            self.id = id
            self.description = description
            self.size_strength = size_strength
            self.manufacturer = manufacturer
            self.quantity = quantity
            if bidDayUnitPrice is None:
                self.bidDayUnitPrice = 0
            else:
                self.bidDayUnitPrice = bidDayUnitPrice

        def equals(self, otherRow: plumbingJob.partRow):
            if (self.id == otherRow.id 
                  and "".join(self.description.split()).lower() == "".join(otherRow.description.split()).lower()
                  and self.size_strength == otherRow.size_strength):
                return True
            return False

        def getIdentity(self) -> str:
            return "{} --> {}".format(self.id, self.description)

        def getTotalPrice(self) -> float:
            if(not (type(self.bidDayUnitPrice) == int or type(self.bidDayUnitPrice) == float)):
                print("{}, {}".format(type(self.bidDayUnitPrice), type(self.quantity)))
            return self.bidDayUnitPrice * self.quantity

        def __str__(self) -> str:
            return "{} {} {} {} {}".format(self.id, self.description, self.quantity, self.size_strength, self.manufacturer)
            #return str(self.job) + " " + str(self.id) + " " + str(self.size_strength) + " " + str(self.manufacturer) + " " + str(self.quantity)

        def __repr__(self) -> str:
            return str(self)
    
    def __init__(self, title, rows: List[partRow] = []):
        self.rows = rows
        self.title = title
    
    def addRow(self, row):
        self.rows.append(row)

    def __str__(self) -> str:
        rowString = ""
        for row in self.rows:
            rowString += str(row) + "\n"
        return rowString

    def __repr__(self) -> str:
        return str(self)


        



finalPath = "Fixtures.xlsx"

final = openpyxl.Workbook()
finalSheet = final.active

def parseFixtures(sourcePath, sourceQuantityCol, sourceStartRow):
    #Update this not to use chr(ord) so much
    source = openpyxl.load_workbook(sourcePath)
    sourceSheet = source['Plbg']
    tagIdCol = chr(ord(sourceQuantityCol) + 1)
    souceManufacturerColumn = chr(ord(tagIdCol) + 1)
    sourceSizeCol = chr(ord(souceManufacturerColumn) + 1)
    identityCol = chr(ord(sourceQuantityCol) + 5)
    finalIdentityCol = 'A'
    descriptionCol = chr(ord(sourceQuantityCol) + 4)
    finalSizeCol = chr(ord(finalIdentityCol) + 1)
    finalManufacturerCol = chr(ord(finalSizeCol) + 1)
    finalQuantityCol = chr(ord(finalManufacturerCol) + 1)
    
    
    finalSheet[sourceQuantityCol + str(sourceStartRow - 1)] = sourcePath
    
    iterator = sourceSheet.iter_rows(min_row=sourceStartRow, max_row=sourceSheet.max_row, max_col=sourceSheet.max_column)
    i = 3
    thisJob = plumbingJob(os.path.basename(sourcePath), [])
    #maybe create a data structure that is a row 
    for row in iterator:
        # if(row[0].value is None):
        #     currentRow = 0
        # else:
        currentRow = row[0].row
        currentRow = str(currentRow)
        currentCell = sourceSheet[sourceQuantityCol + currentRow].value
        if currentCell is not None and currentCell != 0:
            tagID = sourceSheet[tagIdCol + currentRow].value
            quantity = sourceSheet[sourceQuantityCol + currentRow].internal_value
            if "=" in str(quantity):
                quantity = "ERROR" 
            description = sourceSheet[descriptionCol + currentRow].value
            manufacturer = sourceSheet[souceManufacturerColumn + currentRow].value
            size = sourceSheet[sourceSizeCol + currentRow].value
            bidDayUnitPrice = 0
            for c in row:
                if type(c.value) is str and 'SELECTED' in c.value:
                    bidDayUnitPrice = sourceSheet.cell(row=c.row, column=c.column-1).value
            #Find the price by finding the price next to SELECTED in the same row
            #bidDayPrice = sourceSheet[bidDayPriceCol + currentRow].value
            #finalSheet[priceCol + str(i)].value = price
            thisJob.addRow(plumbingJob.partRow(id=tagID, quantity=quantity, size_strength=size, description=description, manufacturer=manufacturer, bidDayUnitPrice=bidDayUnitPrice))
            i += 1
    return thisJob
my_red = openpyxl.styles.colors.Color(rgb='00FF0000')
my_fill = openpyxl.styles.fills.PatternFill(patternType='solid', fgColor=my_red)

if __name__ == '__main__':
    tkinter.messagebox.showinfo("Instructions", "Put your job recap files into a single folder \n you will be prompted to select that folder next")
    sPath = filedialog.askdirectory(title = "Select the directory containing recaps", initialdir = "../")
    #sPath = os.path.basename()
    for f in os.listdir(sPath):
        xl.convert(sPath + "/" + f, "Plbg")

    print(sPath)
    print("Processing...")

    myJobs: List[plumbingJob] = []
    sourcePaths = []
    for f in os.listdir(sPath):
        sourcePaths.append('{}/'.format(sPath) + f)
    sys.argv = ['', 'B', 'Y', 18]
    sourceQuantityCol = sys.argv[1]
    bidDayPriceCol = sys.argv[2]
    sourceStartRow = int(sys.argv[3])
    
    for source in sourcePaths:
        myJobs.append(parseFixtures(source, sourceQuantityCol, sourceStartRow=sourceStartRow))
    rowCounter = 3
    spacing = 3
    rowIndex = {}
    finalSheet['A2'] = "Identification"
    finalSheet['B2'] = "Size/Strength"
    for i, j in enumerate(myJobs):
        baseColumn = (i + len(sourcePaths)) + i*spacing + 1
        finalSheet.cell(column=baseColumn, row=1).value = j.title
        print(j.title)
        finalSheet.cell(column=baseColumn, row=2).value = 'Manufacturer'
        finalSheet.cell(column=baseColumn+1, row=2).value = 'Quantity'
        finalSheet.cell(column=baseColumn+2, row=2).value = 'Bid Day Total Price'
        finalSheet.cell(column=baseColumn+3, row=2).value = 'Bid Day Unit Price'
        for row in j.rows:
            
            #may need to compare rows in a less exact way, so that similar rows are not duplicated
            for r in rowIndex:
                if row.equals(r):
                    #print('Same row found')
                    #print("Manufacturer: {}, Bid Day Unit Price: {}".format(row.manufacturer, row.bidDayUnitPrice))
                    finalSheet.cell(column=baseColumn, row=rowIndex[r]).value = row.manufacturer
                    if("ERROR" in str(row.quantity)):
                        finalSheet.cell(column=baseColumn+1, row=rowIndex[r]).fill = my_fill
                    finalSheet.cell(column=baseColumn+1, row=rowIndex[r]).value = row.quantity
                    finalSheet.cell(column=baseColumn+2, row=rowIndex[r]).value = row.getTotalPrice()
                    finalSheet.cell(column=baseColumn+3, row=rowIndex[r]).value = row.bidDayUnitPrice
                    break
            else:
                row.row = rowCounter
                rowIndex[row] = rowCounter
                rowCounter += 1
                finalSheet.cell(column=1, row=row.row).value = row.getIdentity()
                finalSheet.cell(column=2, row=row.row).value = row.size_strength

                finalSheet.cell(column=baseColumn, row=row.row).value = row.manufacturer

                if("ERROR" in str(row.quantity)):
                    finalSheet.cell(column=baseColumn+1, row=row.row).fill = my_fill
                finalSheet.cell(column=baseColumn+1, row=row.row).value = row.quantity
                finalSheet.cell(column=baseColumn+2, row=row.row).value = row.getTotalPrice()
                finalSheet.cell(column=baseColumn+3, row=row.row).value = row.bidDayUnitPrice                
    #print(myJobs)
    #print(*rowIndex, sep='\n')
    os.chdir('../')
    final.save(finalPath)
    subprocess.run("explorer {}".format(finalPath))
    #fd = filedialog.askopenfile(title="Select Spec File", filetypes=[("PDFs", ".pdf")])
    #print('test')
    