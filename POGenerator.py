from pathlib import Path
import tkinter
from tkinter import ttk
import openpyxl
from tkinter import Entry, Label, Tk, filedialog
import os
import sqlite3
import subprocess
import csv
import FolderSearcher
#import pyexcel as p
#pyexcel doesn't work with nuitka compiler for exe. 
#pyexcel only needed if using xls files
import shutil
from sqlescapy import sqlescape
#Figure out how to set the current working directory to the shortcut location

#things like 2-2 and 2/2 are getting auto converted to dates, need to prevent this! May need to use .xlsx instead of .csv
# "=""2008-10-03""" adding equals sign before the text might work, let's see

#How to speed up startup when running as an EXE?

#Potentially add iterator to make it easier to get attributes (usefull for the query checking with sqlescape) 
class LineItem:
    def __init__(self, type, material, name, size, quantity, job):
        self.type = type
        self.material = material
        self.name = name
        self.size = size
        self.quantity = quantity
        self.job = job

    def __str__(self):
        return self.__repr__()
    def __repr__(self):
        return "{},{},{},{},{},{}\n".format(self.type, self.material, self.name, self.size, self.quantity, self.job)

# Query to Combine rows: Not sure what the having count does
# SELECT SUM(quantity) as 'Total Quantity', Type, Material, Name, Size 
# FROM LineItems 
# GROUP BY Size, Name, Material, Type 
# HAVING COUNT(*)>1;

#Do I need to include labor cost?

jobBoxes = []
jobIndex = 3
submitBtn = None
def deleteBox(box, btn, a = None, b = None ,c = None):
    jobBoxes.remove(box)
    box.destroy()
    btn.destroy()
    if(submitBtn is not None and len(jobBoxes) == 0):
            submitBtn['state'] = tkinter.DISABLED
    
        

def test(box, a = None, b = None ,c = None):
    pass

def submit(a = None, b = None ,c = None):
    jobIDs = []
    for box in jobBoxes:
        jobIDs.append(box.get())
    filePaths = FolderSearcher.getXlFiles(jobIDs)
    print(filePaths)
    makeDatabase(filePaths)

#how to keep these all together?
def createButton(btn):
    if(submitBtn is not None):
            submitBtn['state'] = tkinter.NORMAL
    global jobIndex 
    SV = tkinter.StringVar()
    SV.trace_add("write", callback=test)
    idB = Entry(win, textvariable=SV, font=('Aerial 16'), width=8)
    jobBoxes.append(idB)
    idB.grid(row=jobIndex, column='0', sticky='n')
    lIn = ttk.Button(win, text="Delete", command=lambda: deleteBox(idB, lIn))
    lIn.grid(row=jobIndex, column = 2)
    
    btn.grid(row=jobIndex+1, column=0, sticky='n')
    jobIndex += 1

def convertAndCache(filePath):
    # if(filePath.name.endswith(".xls")):
    #     p.save_book_as(file_name=str(filePath),
    #             dest_file_name='FileCache/' + filePath.stem + '.xlsx')
    if(filePath.name.endswith(".xlsx")):
        #this is giving errors, I think I need to use an absolute path instead of a relative path
        shutil.copy2(filePath, os.getcwd() + "/FileCache")

def cleanString(string):
    return string.replace(" ", "")

def makeDatabase(paths):
    #Did formatting change?
    con = sqlite3.connect('po.db')
    cur = con.cursor()
    #main lives in the same namespace as top level code interesting, so no need for global calls
    #folderPath = filedialog.askdirectory(title="Select Folder With Material Charts")
    #folderPath = "C:\Automation\TakeOffReader\Takeoffs"
    print(cur.execute("DROP TABLE IF EXISTS LineItems").fetchall())
    print(cur.execute("DROP TABLE IF EXISTS Condensed").fetchall())
    print(cur.execute("""CREATE TABLE LineItems(
            Type VARCHAR(50) NOT NULL,
            Material VARCHAR(50) NOT NULL,
            Name VARCHAR(50) NOT NULL,
            Size VARCHAR(50) NOT NULL,
            Quantity REAL NOT NULL,
            Job VARCHAR(50) NOT NULL
            );""").fetchall())
    for file in paths:
        convertAndCache(file)
    #Needs a try catch when opening the files, or a with statement. 
    #To prevent corrupted or incorrectly labeled xl files from crashing the program
    for file in Path("FileCache").glob("*"):
    #     print(p)
    # for file in paths:
        #Use this to convert from xls to xlsx
        # p.save_book_as(file_name=file,
        #        dest_file_name='your-new-file-out.xlsx')
        sourcePath = file
        #sourcePath  = os.path.join(folderPath, file)
        print(sourcePath)
        source = openpyxl.load_workbook(sourcePath)
        sourceSheet = source.active
        rowIter = sourceSheet.iter_rows(max_row=sourceSheet.max_row, max_col=sourceSheet.max_column)
        currentType = "Unknown"
        currentMaterial = "Unknown"
        currentName = "Unknown"
        lineItems = []
        for i, row in enumerate(rowIter):
            if(sourceSheet["A" + str(i+1)].value is not None):
                currentType = sourceSheet["A" + str(i+1)].value
            if(sourceSheet["G" + str(i+1)].value is not None):
                currentMaterial = sourceSheet["G" + str(i+1)].value
            if(sourceSheet["F" + str(i+1)].value is not None and sourceSheet["F" + str(i+1)].value != "Item"):
                currentName = sourceSheet["F" + str(i+1)].value
            #This if may need to be improved, more detailed, it may have exceptions I haven't covered
            if(sourceSheet["I" + str(i+1)].value is not None and sourceSheet["O" + str(i+1)].value is not None):
                lineItems.append(LineItem(currentType, currentMaterial, currentName, sourceSheet["I" + str(i+1)].value, sourceSheet["O" + str(i+1)].value, file.stem))   
        #Should use the csv writer instead of this manual method
        #fileName = file.stem + ".csv"
        #f = open(fileName, "w")
        #f.write("{},{},{},{},{},{}".format("Type", "Material", "Name", "Size", "Quantity","job"))           
        for l in lineItems:
            #f.write(str(l))
            #This is quite repetitive, might benefit slightly from a function
            #Could remove the print statements, after testing
            oldType = str(l.type)
            oldMaterial = str(l.material)
            oldName = str(l.name)
            oldSize = str(l.size)
            oldQuantity = str(l.quantity)
            oldJob = str(l.job)

            l.type = sqlescape(str(l.type))
            if(oldType != str(l.type)):
                print("Type changed from " + oldType + " to " + str(l.Type))
            l.material = sqlescape(str(l.material))
            if(oldMaterial != str(l.material)):
                print("Material changed from " + oldMaterial + " to " + str(l.material))
            l.name = sqlescape(str(l.name))
            if(oldName != str(l.name)):
                print("Name changed from " + oldName + " to " + str(l.name))
            l.size = sqlescape(str(l.size))
            if(oldSize != str(l.size)):
                print("Size changed from " + oldSize + " to " + str(l.size))
            l.quantity = sqlescape(str(l.quantity))
            if(oldQuantity != str(l.quantity)):
                print("Quantity changed from " + oldQuantity + " to " + str(l.quantity))
            l.job = sqlescape(str(l.job))
            if(oldJob != str(l.job)):
                print("Job changed from " + oldJob + " to " + str(l.job))
            
            #Adding an equal sign or quotessomething like that to each value would prevent it from being converted into dates
            # '=' + str(l.size) something like this, but = doesn't work when the input is not a number
            # Maybe I should clean the input files instead of the output
            # The single quote method is alright, even though the output has quotes in it
            query = 'INSERT INTO LineItems(Type, Material, Name, Size, Quantity, Job) VALUES("{}", "{}", "{}", "\'{}\'", {}, "{}")'.format(l.type, l.material, l.name, l.size, l.quantity, l.job)
            try:
                cur.execute(query)
            except:
                print("Error writing to database. Failed input below:")
                print(l)
            #print(query)
            
    #if(fileName is not None):
        #f.close()
    #May want # HAVING COUNT(*)>1; at the end I don't know what this does though
    #Create a separate table with this data in the database:
    #New table for each job maybe?
    #Changing directories this way is slightly dumb, but it works
    os.chdir("../")
    with open('MaterialsOutput.csv', 'w', newline='') as f:
        writer = csv.writer(f)
        writer.writerow(['Type', 'Material', 'Name', 'Size', 'Quantity', 'Job'])
        writer.writerow(['', '', '', '', '', 'All Jobs Combined'])
        cur.execute("""CREATE TABLE Condensed AS SELECT Type, Material, Name, Size, SUM(quantity) as 'Total Quantity' 
        FROM LineItems GROUP BY Size, Name, Material, Type ORDER BY Type asc, Material asc, Name asc, Size asc""")
        conData = cur.execute("SELECT * FROM Condensed").fetchall()
        print(*conData, sep='\n')
        writer.writerows(conData)
        writer.writerow(['', '', '', '', '', '', ''])
        lineData = cur.execute("SELECT * FROM LineItems").fetchall()
        print(*lineData, sep='\n')
        writer.writerows(lineData)
        
    
    con.commit()
    subprocess.run("explorer MaterialsOutput.csv")
if __name__ == '__main__':
    while(True):
        try:
            f = open('MaterialsOutput.csv', 'w', newline='')
            f.close()
            break
        except:
            tkinter.messagebox.showinfo("Error with MaterialsOutput.csv","Please close MaterialsOutput.csv and try again")
    print(os.getcwd())
    if(os.path.exists("FileCache")):
        shutil.rmtree("FileCache")
    os.mkdir("FileCache")
    win = Tk()
    win.geometry("500x500")
    ws = win.winfo_screenwidth() # width of the screen
    hs = win.winfo_screenheight() # height of the screen

    # calculate x and y coordinates for the Tk root window
    x = (ws/2) - 150
    y = (hs/2) - 200
    win.geometry('%dx%d+%d+%d' % (200, 200, x, y))
    win.geometry("")
    win.minsize(width=200, height=300)
    win.resizable(True,True)
    win.title("Purchase Order Generator")
    Label(win, text="Estimating Job Numbers ####:", font=('Aerial 16 bold')).grid(row='0', column='0')
    #plus button that adds another input box
    
    addBtn = ttk.Button(win, text="Add", command= lambda: createButton(addBtn))
    addBtn.grid(row=1, column = 0, sticky='n')
    submitBtn = ttk.Button(win, text="Submit", command=submit, state=tkinter.DISABLED)
    submitBtn.grid(row=100, column = 0)
    win.mainloop()
    # print(xlsFiles)
    
    #print(*(os.listdir(folderPath)), sep='\n')

    #for file in os.listdir(folderPath)
    

        
    
    
    
    
    