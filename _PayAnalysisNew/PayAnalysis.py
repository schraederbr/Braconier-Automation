import openpyxl
import sys
import dateparser
from datetime import datetime
from datetime import timedelta
from openpyxl.descriptors.base import DateTime


SOURCEPATH = "SageData.xlsx"
TIMECARDPATH = "Timecards Procore.xlsx"
TRACKINGSHEET = "Beth Tracking Sheet.xlsx"
FINALPATH = "final.xlsx"

class Employee:
    def __init__(self, id, name = "UNKNOWN NAME", regHours = 0, otHours = 0, regRate = 0, otRate = 0, date: datetime = datetime(1, 1, 1)):
            self.date = date
            self.id = id
            self.name = name
            self.regHours = regHours
            self.otHours = otHours
            self.regRate = regRate
            self.otRate = otRate
    
    def __str__(self) -> str:
        return "Employee(date = {}, id = {}, name = {}, regHours = {}, otHours = {}, payRate = {}, otRate = {}, totalPay = {}) \n".format(self.date, self.id, self.name, self.regHours, self.otHours, self.regRate, self.otRate, self.totalPay())

    def __repr__(self) -> str:
        return self.__str__()

    def add(self, other):
        self.regHours += other.regHours
        self.otHours += other.otHours
        if self.regRate != other.regRate and self.regRate != 0 and other.regRate != 0:
            print("Error: regRate mismatch")
            print("sr: {} or: {}".format(self.regRate, other.regRate))
        if self.otRate != other.otRate and self.otRate != 0 and other.otRate != 0:
            print("Error: otRate mismatch")
            print("sot: {} ort: {}".format(self.otRate, other.otRate))
        self.regRate = other.regRate if self.regRate == 0 else self.regRate
        self.otRate = other.otRate if self.otRate == 0 else self.otRate

        return self
    
    def equalPay(self, other):
        return self.regRate == other.regRate and self.otRate == other.otRate and self.regHours == other.regHours and self.otHours == other.otHours

    def totalPay(self):
        return self.regHours * self.regRate + self.otHours * self.otRate


# class Crew:
#     def __init__(self, id, employees: list[Employee] = []):
#         self.id = id
#         self.employees = employees

class PayPeriod:
    def __init__(self, id, employees: list[Employee] = []):
        self.id = id
        self.employees = employees
        
    def totalHours(self) -> float:
        return sum([employee.regHours + employee.otHours for employee in self.employees])

    def regularHours(self) -> float:
        return sum([employee.regHours for employee in self.employees])

    def overtimeHours(self) -> float:
        return sum([employee.otHours for employee in self.employees])

    def totalPay(self):
        total = 0
        for e in self.employees:
            reg = e.regHours * e.regRate
            ot = e.otHours * e.otRate
            total += reg + ot
        return total

    def __str__(self) -> str:
        #Could shorten these long return lines 
        return "Date: {} totalHours: {} regularHours: {} overTimeHours {} Total Pay: {} \n {}".format(self.id,
          self.totalHours(), self.regularHours(), self.overtimeHours(), self.totalPay(), self.employees)

    def __repr__(self) -> str:
        return self.__str__()

# payrates by last name
payRates = {
    

}

if __name__ == '__main__':
    unknownEmps = [] 
    trackSheetName = "Hours Totals Wk15"
    if len(sys.argv) > 1:
        trackSheetName = sys.argv[1]
    timecard = openpyxl.load_workbook(TIMECARDPATH)
    timecardSheet = timecard.active
    rowIter = timecardSheet.iter_rows(max_row=timecardSheet.max_row, max_col=timecardSheet.max_column, min_row=3)    
    seen = []
    names = {}
    for row in rowIter:
        if(row[1].value not in seen):
            seen.append(row[1].value)
            names[row[2].value] = row[1].value
    #print(names)
    payPeriods = []
    for i in range(1, 19):
        trackSheetName = "Hours Totals Wk{}".format(i)
        tracking = openpyxl.load_workbook(TRACKINGSHEET)
        #print(tracking.sheetnames)
        trackingSheet = tracking[trackSheetName]
        rowIter = trackingSheet.iter_rows(max_row=trackingSheet.max_row, max_col=trackingSheet.max_column, min_row=3)
        payPeriod = PayPeriod('2000-01-01 00:00:00')
        invNames = {v.strip(): k.strip() for k, v in names.items()}
        lastNames = {v.split(' ')[1].strip(): k for v, k in invNames.items()}
        for row in rowIter:
            try: 
                type(int(row[2].value))
                if(row[0].value is not None and row[0].value != 0 and row[2].value != 0 and row[1].value is not None and not "Crew" in row[1].value):
                #FFFFFF00 is yellow
                #print('RGB =', tuple(int(color_in_hex[i:i+2], 16) for i in (0, 2, 4))) # Color in RGB
                    if "," in row[1].value:
                        #print("Broken name: " + row[1].value)
                        fixedName = (row[1].value.split(",")[1].strip() + " " + row[1].value.split(",")[0]).strip()
                        #print(fixedName)
                        row[1].value = fixedName
                    if row[1].value in invNames:
                        employee = Employee(invNames[row[1].value], row[1].value, date=row[0].value)
                    elif row[1].value.split(' ')[1] in lastNames:
                        employee = Employee(lastNames[row[1].value.split(' ')[1].strip()], row[1].value, date=row[0].value)
                    else:
                        #Ignore names if they don't match
                        employee = Employee("NAME NOT FOUND", row[1].value, date=row[0].value)
                    hexColor = row[2].fill.start_color.index
                    
                    #this doesn't work well if overtime cell color is not changed
                    if hexColor is not None and hexColor != "00000000" and hexColor != "0" and hexColor != 0:
                        employee.otRate = row[4].value
                        employee.otHours = float(row[2].value)
                        if(hexColor != "FFFFFF00"):
                            print(hexColor)
                            print(employee.name + "COLOR CHANGE DETECTED, CONFIRM")
                            
                    else:
                        employee.regRate = row[4].value
                        employee.regHours = float(row[2].value)
                    if(employee.id != "NAME NOT FOUND"):
                        payPeriod.employees.append(employee)
                    else:
                        unknownEmps.append(employee)
                
                date = datetime.strptime(str(row[0].value), '%Y-%m-%d %H:%M:%S')
                weekDate = datetime.strptime(str(payPeriod.id), '%Y-%m-%d %H:%M:%S')
                if(date.date() > weekDate.date()):
                    payPeriod.id = date
            except:
                pass
        payPeriods.append(payPeriod)
        #except Exception as e: print(repr(e))
    #print(str(week))
    #print("\n\n\n\n\n\n\n\n")
    #print(week)
    #print(week.totalPay())
    
    #sortedPeriod = PayPeriod("test", sorted(payPeriod.employees, key=lambda x: x.date))
    #print(sortedPeriod)


    emps = []  
    current = None
    for e in payPeriod.employees:
        if current is None:
            current = e
            continue
        if current.id == "NAME NOT FOUND":
            unknownEmps.append(current)
            current = e
            continue
        if current.id != e.id or current.name != e.name:
            emps.append(current)
            current = e
        else:
            current = current.add(e)
    #This doesn't quite seem to be working
    if current is not None and current.id != "NAME NOT FOUND":
        emps.append(current)
    #print(emps)
    myWe = PayPeriod(payPeriod.id, emps)
    #print(myWe)

    output = ""
    for e in unknownEmps:
        if not e.name + " " in output :
            output += e.name + " "
    #print("Unknown Employees \n" + output)
    #print(newWeek.totalPay())



    final = openpyxl.Workbook()
    source = openpyxl.load_workbook(SOURCEPATH)
    sourceSheet = source.active
    finalSheet = final.active
    currentDate = "NOTHING"
    rowIter = sourceSheet.iter_rows(max_row=sourceSheet.max_row, max_col=sourceSheet.max_column, min_row=3)    
    
    payWeeks = []
    currentPayPeriod = None
    currentEmployee = None
    for row in rowIter:
        if(row is not None and row[5].value is not None):
            if currentEmployee is not None:
                if currentPayPeriod is not None:
                    if currentEmployee.id != row[2].value.strip():
                        currentPayPeriod.employees.append(currentEmployee)
                        if(row[2].value.strip() in names):
                            currentEmployee = Employee(row[2].value.strip(), names[row[2].value.strip()])
                        else:
                            currentEmployee = Employee(row[2].value.strip())
            else:
                currentEmployee = Employee(row[2].value.strip())
            if '1' in row[5].value:
                currentEmployee.regHours += float(row[6].value)
            elif '2' in row[5].value:
                currentEmployee.otHours += float(row[6].value)
            if row[0].value != currentDate:
                currentPayPeriod = PayPeriod(row[0].value, [])
                payWeeks.append(currentPayPeriod)
                currentDate = row[0].value

    currentPayPeriod.employees.append(currentEmployee)
    #print(payWeeks)
    total = 0.0
    for payPeriod in payWeeks:
        total += payPeriod.totalHours()
    #print(payWeeks[1])

    #get dates of all the pay periods
    dates  = []
    for payPeriod in payWeeks:
        #print(payPeriod.id)
        dates.append(datetime.strptime(str(payPeriod.id).strip(), '%m-%d-%Y'))
    dates = sorted(dates)
    dateBins = []
    dateBins.append((dates[0] - timedelta(days=7), dates[0], PayPeriod(dates[0].strftime('%m-%d-%Y'), [])))
    for i in range(1, len(dates)):
        dateBins.append((dates[i-1] + timedelta(days=1), dates[i], PayPeriod(dates[i].strftime('%m-%d-%Y'), [])))
    
    
    for e in payPeriod.employees:
        for b in dateBins:
            if b[0] <= e.date <= b[1]:
                b[2].employees.append(e) 
    for b in dateBins:
        b[2].employees = sorted(b[2].employees, key=lambda x: x.date)
        binTotal = 0
        for e in b[2].employees:
            binTotal += e.totalPay()
        #print(binTotal)
        print(b)
    
    
    
    



