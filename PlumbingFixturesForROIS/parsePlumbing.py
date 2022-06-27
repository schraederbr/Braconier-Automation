from __future__ import annotations
import os
import subprocess
import sys
from typing import List

import openpyxl
from tkinter import filedialog
import pycel





class plumbingJob:
    class partRow:
        def __init__(self, id, description: str, quantity, size_strength = None
          , manufacturer = "unknown", row = None, bidDayUnitPrice = None):
            self.row = row
            self.id = id
            self.description = description
            self.size_strength = size_strength
            self.manufacturer = manufacturer
            self.quantity = quantity
            if bidDayUnitPrice is None:
                self.bidDayUnitPrice = 0
            else:
                self.bidDayUnitPrice = bidDayUnitPrice

        def equals(self, otherRow: plumbingJob.partRow):
            if (self.id == otherRow.id 
                  and "".join(self.description.split()) == "".join(otherRow.description.split())
                  and self.size_strength == otherRow.size_strength):
                return True
            return False

        def getIdentity(self) -> str:
            return "{} --> {}".format(self.id, self.description)

        def getTotalPrice(self) -> float:
            return self.bidDayUnitPrice * self.quantity

        def __str__(self) -> str:
            return "{} {} {} {} {}".format(self.id, self.description, self.quantity, self.size_strength, self.manufacturer)
            #return str(self.job) + " " + str(self.id) + " " + str(self.size_strength) + " " + str(self.manufacturer) + " " + str(self.quantity)

        def __repr__(self) -> str:
            return str(self)
    
    def __init__(self, title, rows: List[partRow] = [], column = 'A'):
        self.rows = rows
        self.title = title
        self.column = column
    
    def addRow(self, row):
        self.rows.append(row)

    def __str__(self) -> str:
        rowString = ""
        for row in self.rows:
            rowString += str(row) + "\n"
        return rowString

    def __repr__(self) -> str:
        return str(self)


        



finalPath = "FixtureTakeoff.xlsx"

final = openpyxl.Workbook()
finalSheet = final.active

def parseFixtures(sourcePath, sourceQuantityCol, sourceStartRow):
    source = openpyxl.load_workbook(sourcePath)
    sourceSheet = source['Plbg']
    tagIdCol = chr(ord(sourceQuantityCol) + 1)
    souceManufacturerColumn = chr(ord(tagIdCol) + 1)
    sourceSizeCol = chr(ord(souceManufacturerColumn) + 1)
    identityCol = chr(ord(sourceQuantityCol) + 5)
    finalIdentityCol = 'A'
    descriptionCol = chr(ord(sourceQuantityCol) + 4)
    finalSizeCol = chr(ord(finalIdentityCol) + 1)
    finalManufacturerCol = chr(ord(finalSizeCol) + 1)
    finalQuantityCol = chr(ord(finalManufacturerCol) + 1)
    
    
    finalSheet[sourceQuantityCol + str(sourceStartRow - 1)] = sourcePath
    
    iterator = sourceSheet.iter_rows(min_row=sourceStartRow, max_row=sourceSheet.max_row, max_col=sourceSheet.max_column)
    #finalSheet[identityCol + '1'] = "Identification"
    #finalSheet[finalSizeCol + '1'] = "Size/Strength"
    #finalSheet[finalManufacturerCol + '1'] = "Manufacturer"
    #finalSheet[finalQuantityCol + '1'] = "Original Quantity Per Fixture"
    i = 3
    thisJob = plumbingJob(sourcePath.split('.')[0].split('/')[1], [])
    #maybe create a data structure that is a row 
    for row in iterator:
        currentRow = row[0].row
        currentRow = str(currentRow)
        currentCell = sourceSheet[sourceQuantityCol + currentRow].value
        if currentCell is not None and currentCell != 0:
            tagID = sourceSheet[tagIdCol + currentRow].value
            #finalSheet[tagIdCol + str(i)].value = tagID
            quantity = sourceSheet[sourceQuantityCol + currentRow].internal_value
            if "=" in str(quantity):
                quantity = "ERROR" 
            #finalSheet[finalQuantityCol + str(i)].value = quantity
            description = sourceSheet[descriptionCol + currentRow].value
            #finalSheet[descriptionCol + str(i)].value = description
            identification = "{} --> {}".format(tagID, description)
            #finalSheet[finalIdentityCol + str(i)].value = identification
            manufacturer = sourceSheet[souceManufacturerColumn + currentRow].value
            #finalSheet[finalManufacturerCol + str(i)].value = manufacturer
            size = sourceSheet[sourceSizeCol + currentRow].value
            #finalSheet[finalSizeCol + str(i)].value = size
            bidDayUnitPrice = 0
            for c in row:
                if type(c.value) is str and 'SELECTED' in c.value:
                    bidDayUnitPrice = sourceSheet.cell(row=c.row, column=c.column-1).value
            #Find the price by finding the price next to SELECTED in the same row
            #bidDayPrice = sourceSheet[bidDayPriceCol + currentRow].value
            #finalSheet[priceCol + str(i)].value = price
            thisJob.addRow(plumbingJob.partRow(id=tagID, quantity=quantity, size_strength=size, description=description, manufacturer=manufacturer, bidDayUnitPrice=bidDayUnitPrice))
            i += 1
    return thisJob
my_red = openpyxl.styles.colors.Color(rgb='00FF0000')
my_fill = openpyxl.styles.fills.PatternFill(patternType='solid', fgColor=my_red)

if __name__ == '__main__':
    # I wonder if there is a way to without all the uses of chr(ord()) like a way to increment the row and column
    # using chr(ord(column) + 1) fails if you overflow from Z, it doesn't give you AA
    # maybe like this: bidDayUnitPrice = sourceSheet.cell(row=c.row, column=c.column).value

    myJobs: List[plumbingJob] = []
    sourcePaths = []
    for f in os.listdir('jobs'):
        sourcePaths.append('jobs/' + f)
    sys.argv = ['', 'B', 'Y', 18]
    sourceQuantityCol = sys.argv[1]
    bidDayPriceCol = sys.argv[2]
    sourceStartRow = int(sys.argv[3])
    
    for source in sourcePaths:
        myJobs.append(parseFixtures(source, sourceQuantityCol, sourceStartRow=sourceStartRow))
    rowCounter = 3
    spacing = 3
    rowIndex = {}
    #myJo#bs[0].addRow(plumbingJob.partRow(id="100", quantity="10", size_strength="2'", description="Nothing", manufacturer="Test"))
    #myJobs[1].addRow(plumbingJob.partRow(id="100", quantity="10", size_strength="2'", description="Nothing", manufacturer="Test"))
    #myJobs[2].addRow(plumbingJob.partRow(id="100", quantity="10", size_strength="2'", description="Nothing", manufacturer="Test"))
    finalSheet['A2'] = "Identification"
    finalSheet['B2'] = "Size/Strength"
    for i, j in enumerate(myJobs):

        j.column = chr(ord('A') + (i + len(sourcePaths)))
        finalSheet[chr(ord(j.column)+i*spacing) + str(1)] = j.title
        finalSheet[chr(ord(j.column)+i*spacing) + str(2)] = 'Manufacturer'
        finalSheet[chr(ord(j.column)+i*spacing+1) + str(2)] = 'Quantity'
        finalSheet[chr(ord(j.column)+i*spacing+2) + str(2)] = 'Bid Day Total Price'
        finalSheet[chr(ord(j.column)+i*spacing+3) + str(2)] = 'Bid Day Unit Price'
        for row in j.rows:
            #may need to compare rows in a less exact way, so that similar rows are not duplicated
            for r in rowIndex:
                if row.equals(r):
                    print('Same row found')
                    #finalSheet[j.column + str(rowIndex[r])] = str(row)
                    finalSheet[chr(ord(j.column)+i*spacing) + str(rowIndex[r])] = row.manufacturer
                    if("ERROR" in str(row.quantity)):
                        finalSheet[chr(ord(j.column)+i*spacing+1) + str(rowIndex[r])].fill = my_fill
                    finalSheet[chr(ord(j.column)+i*spacing+1) + str(rowIndex[r])] = row.quantity
                    finalSheet[chr(ord(j.column)+i*spacing+2) + str(rowIndex[r])] = row.getTotalPrice()
                    finalSheet[chr(ord(j.column)+i*spacing+3) + str(rowIndex[r])] = row.bidDayUnitPrice
                    break
            else:
                row.row = rowCounter
                rowIndex[row] = rowCounter
                rowCounter += 1
                finalSheet['A' + str(row.row)] = row.getIdentity()
                finalSheet['B' + str(row.row)] = row.size_strength
                #finalSheet['C' + str(row.row)] = row.manufacturer
                finalSheet[chr(ord(j.column)+i*spacing) + str(row.row)] = row.manufacturer
                if("ERROR" in str(row.quantity)):
                    finalSheet[chr(ord(j.column)+i*spacing+1) + str(row.row)].fill = my_fill
                finalSheet[chr(ord(j.column)+i*spacing+1) + str(row.row)] = row.quantity
                finalSheet[chr(ord(j.column)+i*spacing+2) + str(row.row)] = row.getTotalPrice()
                finalSheet[chr(ord(j.column)+i*spacing+3) + str(row.row)] = row.bidDayUnitPrice
                #finalSheet[chr(ord(j.column) + 1) + str(row.row)] = str(row)
                
    #print(myJobs)
    #print(*rowIndex, sep='\n')
    final.save(finalPath)
    subprocess.run("explorer {}".format(finalPath))
    #fd = filedialog.askopenfile(title="Select Spec File", filetypes=[("PDFs", ".pdf")])
    #print('test')
    